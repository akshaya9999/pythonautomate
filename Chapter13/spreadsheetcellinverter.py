import openpyxl

wb=openpyxl.load_workbook('texttospreadsheet.xlsx')
sheet=wb.active
rowno=sheet.max_row
columnno=sheet.max_column

wbb=openpyxl.Workbook()
invsheet=wbb.active


for i in range(1,rowno+1):
    for j in range(1,columnno+1):
        invsheet.cell(row=j,column=i).value=sheet.cell(row=i,column=j).value
        
wbb.save('invertedspreadsheet.xlsx')
