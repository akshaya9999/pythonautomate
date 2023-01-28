import openpyxl

wb=openpyxl.load_workbook('texttospreadsheet.xlsx')
sheet=wb.active

nolines=sheet.max_row
nofiles=sheet.max_column


for i in range (1,nofiles+1):
    filen='exeltext'+str(i)+'.txt'
    fileop=open(filen,'w')
    for j in range(1,nolines+1):
        fileop.write(sheet.cell(row=j,column=i).value)
    fileop.close()