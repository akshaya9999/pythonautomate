import os
import openpyxl
import csv

for excelFile in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue
    wb=openpyxl.load_workbook(excelFile)
    # Loop through every sheet in the workbook.
    for sheetName in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(sheetName)
        csvFile=open(excelFile[:-5]+'_'+sheetName+'.csv','w')
        csvwriter=csv.writer(csvFile)

        # Create the CSV filename from the Excel filename and sheet title.
        # Create the csv.writer object for this CSV file.

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(sheet.cell(row=rowNum,column=colNum).value)

            # Write the rowData list to the CSV file.
            for row in rowData:
                csvwriter.writerow(row)

        csvFile.close()